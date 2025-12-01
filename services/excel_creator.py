from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime
from io import BytesIO
from config import settings

class ExcelCreator:
    def __init__(self):
        self.default_font_name = settings.DEFAULT_FONT_NAME
        self.default_font_size = settings.DEFAULT_FONT_SIZE
    
    def create_workbook(self, data: dict, filename: str = None) -> BytesIO:
        """Create an Excel workbook from data and return as BytesIO"""
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Process data and create sheets
        for sheet_name, sheet_data in data.items():
            self.create_sheet(wb, sheet_name, sheet_data)
        
        # Save to BytesIO
        wb_stream = BytesIO()
        wb.save(wb_stream)
        wb_stream.seek(0)
        
        return wb_stream
    
    def create_sheet(self, wb, sheet_name, sheet_data):
        """Create a sheet with data"""
        # Create sheet
        ws = wb.create_sheet(title=sheet_name)
        
        # Process sheet data
        if isinstance(sheet_data, dict):
            if 'headers' in sheet_data and 'rows' in sheet_data:
                # Tabular data
                self.create_tabular_data(ws, sheet_data)
            elif 'charts' in sheet_data:
                # Charts
                self.create_charts(ws, sheet_data)
            else:
                # Key-value pairs
                self.create_key_value_data(ws, sheet_data)
        elif isinstance(sheet_data, list):
            # List of rows
            self.create_list_data(ws, sheet_data)
    
    def create_tabular_data(self, ws, sheet_data):
        """Create tabular data with headers and rows"""
        headers = sheet_data.get('headers', [])
        rows = sheet_data.get('rows', [])
        formatting = sheet_data.get('formatting', {})
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            
            # Apply header formatting
            header_format = formatting.get('header', {})
            self.apply_cell_formatting(cell, header_format)
        
        # Add rows
        for row_idx, row in enumerate(rows, 2):
            if isinstance(row, dict):
                # Row as dictionary
                for col_idx, header in enumerate(headers, 1):
                    if header in row:
                        cell = ws.cell(row=row_idx, column=col_idx, value=row[header])
                        
                        # Apply cell formatting
                        cell_format = formatting.get('cells', {}).get(header, {})
                        self.apply_cell_formatting(cell, cell_format)
            else:
                # Row as list
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Apply cell formatting
                    cell_format = formatting.get('cells', {}).get(col_idx-1, {})
                    self.apply_cell_formatting(cell, cell_format)
        
        # Apply column formatting
        for col_idx, col_format in formatting.get('columns', {}).items():
            if isinstance(col_idx, str):
                # Column letter
                col_letter = col_idx
            else:
                # Column index
                col_letter = get_column_letter(col_idx + 1)
            
            # Set column width
            if 'width' in col_format:
                ws.column_dimensions[col_letter].width = col_format['width']
    
    def create_key_value_data(self, ws, sheet_data):
        """Create key-value data"""
        row_idx = 1
        
        for key, value in sheet_data.items():
            # Add key
            key_cell = ws.cell(row=row_idx, column=1, value=key)
            key_cell.font = Font(bold=True)
            
            # Add value
            value_cell = ws.cell(row=row_idx, column=2, value=value)
            
            row_idx += 1
    
    def create_list_data(self, ws, sheet_data):
        """Create list data"""
        for row_idx, row in enumerate(sheet_data, 1):
            if isinstance(row, dict):
                # Row as dictionary
                for col_idx, (key, value) in enumerate(row.items(), 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                # Row as list
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
    
    def create_charts(self, ws, sheet_data):
        """Create charts"""
        charts = sheet_data.get('charts', [])
        
        for chart_config in charts:
            chart_type = chart_config.get('type', 'bar')
            title = chart_config.get('title', 'Chart')
            data_range = chart_config.get('data_range', 'A1:B10')
            position = chart_config.get('position', 'D5')
            
            # Create chart
            if chart_type == 'bar':
                chart = BarChart()
            # Add other chart types as needed
            
            chart.title = title
            
            # Set data range
            data = Reference(ws, range_string=data_range)
            chart.add_data(data)
            
            # Add chart to sheet
            ws.add_chart(chart, position)
    
    def apply_cell_formatting(self, cell, formatting):
        """Apply formatting to a cell"""
        # Font
        if 'font' in formatting:
            font_config = formatting['font']
            font = Font(
                name=font_config.get('name', self.default_font_name),
                size=font_config.get('size', self.default_font_size),
                bold=font_config.get('bold', False),
                italic=font_config.get('italic', False),
                color=font_config.get('color', '000000')
            )
            cell.font = font
        
        # Alignment
        if 'alignment' in formatting:
            alignment_config = formatting['alignment']
            alignment = Alignment(
                horizontal=alignment_config.get('horizontal', 'general'),
                vertical=alignment_config.get('vertical', 'bottom'),
                wrap_text=alignment_config.get('wrap_text', False)
            )
            cell.alignment = alignment
        
        # Border
        if 'border' in formatting:
            border_config = formatting['border']
            side = Side(
                border_style=border_config.get('style', 'thin'),
                color=border_config.get('color', '000000')
            )
            
            border = Border(
                left=side if border_config.get('left', True) else None,
                right=side if border_config.get('right', True) else None,
                top=side if border_config.get('top', True) else None,
                bottom=side if border_config.get('bottom', True) else None
            )
            cell.border = border
        
        # Fill
        if 'fill' in formatting:
            fill_config = formatting['fill']
            fill = PatternFill(
                start_color=fill_config.get('color', 'FFFFFF'),
                end_color=fill_config.get('color', 'FFFFFF'),
                fill_type=fill_config.get('type', 'solid')
            )
            cell.fill = fill
    
    def generate_filename(self, filename: str = None) -> str:
        """Generate a filename with timestamp if not provided"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = filename or f"excel_{timestamp}"
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        return filename
    
    def generate_object_name(self, filename: str) -> str:
        """Generate MinIO object name with date-based folder structure"""
        today = datetime.now()
        return f"{today.strftime('%Y')}/{today.strftime('%m')}/{today.strftime('%d')}/{filename}"