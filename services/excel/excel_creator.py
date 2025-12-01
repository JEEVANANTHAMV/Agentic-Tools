from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import re
from datetime import datetime
from io import BytesIO
from config import settings
import os

class ExcelCreator:
    def __init__(self):
        self.default_font_name = settings.DEFAULT_FONT_NAME
        self.default_font_size = settings.DEFAULT_FONT_SIZE
    
    def create_excel_from_content(self, content: str, filename: str = None) -> BytesIO:
        """Create an Excel workbook from string content and return as BytesIO"""
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb.active)
        
        # Parse content and create sheets
        self.parse_and_format_content(wb, content)
        
        # Save to BytesIO
        excel_stream = BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        
        return excel_stream
    
    def parse_and_format_content(self, wb, content):
        """
        Parse content string and create Excel sheets
        Supports:
        # Sheet Name
        |Header1|Header2|Header3|
        |-------|-------|-------|
        |Row1Col1|Row1Col2|Row1Col3|
        [BOLD]Bold text[/BOLD]
        [ITALIC]Italic text[/ITALIC]
        [BORDER]Text with border[/BORDER]
        [COLOR:RRGGBB]Colored text[/COLOR]
        [ALIGN:left|center|right]Aligned text[/ALIGN]
        """
        lines = content.split('\n')
        i = 0
        current_ws = None
        
        while i < len(lines):
            line = lines[i].strip()
            
            if not line:
                i += 1
                continue
            
            # Check if this is a sheet name (starts with #)
            if line.startswith('#'):
                sheet_name = line.replace('#', '').strip()
                current_ws = wb.create_sheet(title=sheet_name)
                i += 1
                continue
            
            # If no sheet has been created yet, create a default one
            if current_ws is None:
                current_ws = wb.create_sheet(title="Sheet1")
            
            # Check if this is a table
            if line.startswith('|') and '|' in line[1:]:
                # Collect all table lines
                table_lines = [line]
                j = i + 1
                while j < len(lines) and lines[j].strip().startswith('|'):
                    table_lines.append(lines[j].strip())
                    j += 1
                
                # Create table
                self.create_table_from_markdown(current_ws, table_lines)
                i = j
                continue
            
            # Handle regular text
            if current_ws.max_row == 1 and current_ws.max_column == 1 and current_ws.cell(row=1, column=1).value is None:
                # First cell in the sheet
                cell = current_ws.cell(row=1, column=1)
                self.process_cell_formatting(cell, line)
            else:
                # Add a new row
                row = current_ws.max_row + 1
                cell = current_ws.cell(row=row, column=1)
                self.process_cell_formatting(cell, line)
            
            i += 1
    
    def create_table_from_markdown(self, ws, table_lines):
        """
        Create table from markdown syntax
        |Header1|Header2|Header3|
        |-------|-------|-------|
        |Row1Col1|Row1Col2|Row1Col3|
        |Row2Col1|Row2Col2|Row2Col3|
        """
        lines = [line.strip() for line in table_lines if line.strip()]
        
        if not lines:
            return
        
        # Parse headers
        headers = [cell.strip() for cell in lines[0].split('|') if cell.strip()]
        
        # Parse data rows (skip separator line)
        data_rows = []
        for line in lines[2:]:  # Skip header and separator
            cells = [cell.strip() for cell in line.split('|') if cell.strip()]
            if cells:
                data_rows.append(cells)
        
        # Determine starting row
        start_row = ws.max_row + 1 if ws.cell(row=1, column=1).value is not None else 1
        
        # Add headers
        for i, header in enumerate(headers):
            cell = ws.cell(row=start_row, column=i+1)
            cell.value = header
            cell.font = Font(bold=True, name=self.default_font_name, size=self.default_font_size)
            cell.alignment = Alignment(horizontal='center')
            self.set_cell_border(cell)
        
        # Add data rows
        for row_idx, row_data in enumerate(data_rows):
            for col_idx, cell_data in enumerate(row_data):
                cell = ws.cell(row=start_row + row_idx + 1, column=col_idx+1)
                self.process_cell_formatting(cell, cell_data)
                self.set_cell_border(cell)
        
        # Adjust column widths
        for col_idx in range(len(headers)):
            col_letter = get_column_letter(col_idx + 1)
            ws.column_dimensions[col_letter].width = 15
    
    def process_cell_formatting(self, cell, text):
        """Process cell formatting like bold, italic, color, alignment, and borders"""
        # Default formatting
        cell.font = Font(name=self.default_font_name, size=self.default_font_size)
        
        # Process formatting tags
        # Bold
        bold_pattern = r'\[BOLD\](.*?)\[/BOLD\]'
        bold_match = re.search(bold_pattern, text)
        if bold_match:
            text = re.sub(bold_pattern, bold_match.group(1), text)
            cell.font = cell.font.copy(bold=True)
        
        # Italic
        italic_pattern = r'\[ITALIC\](.*?)\[/ITALIC\]'
        italic_match = re.search(italic_pattern, text)
        if italic_match:
            text = re.sub(italic_pattern, italic_match.group(1), text)
            cell.font = cell.font.copy(italic=True)
        
        # Color
        color_pattern = r'\[COLOR:([A-Fa-f0-9]{6})\](.*?)\[/COLOR\]'
        color_match = re.search(color_pattern, text)
        if color_match:
            text = re.sub(color_pattern, color_match.group(2), text)
            cell.font = cell.font.copy(color=color_match.group(1))
        
        # Alignment
        align_pattern = r'\[ALIGN:(left|center|right)\](.*?)\[/ALIGN\]'
        align_match = re.search(align_pattern, text)
        if align_match:
            text = re.sub(align_pattern, align_match.group(2), text)
            alignment = align_match.group(1)
            if alignment == 'center':
                cell.alignment = Alignment(horizontal='center')
            elif alignment == 'right':
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
        
        # Border
        border_pattern = r'\[BORDER\](.*?)\[/BORDER\]'
        border_match = re.search(border_pattern, text)
        if border_match:
            text = re.sub(border_pattern, border_match.group(1), text)
            self.set_cell_border(cell)
        
        # Set the cell value
        cell.value = text
    
    def set_cell_border(self, cell):
        """Set cell borders"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border
    
    def generate_filename(self, filename: str = None) -> str:
        """Generate a filename with timestamp if not provided"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = filename or f"excel_{timestamp}"
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        return filename
    
    def generate_object_name(self, filename: str) -> str:
        """Generate MinIO object name with date-based folder structure"""
        today = datetime.now()
        return f"{today.strftime('%Y')}/{today.strftime('%m')}/{today.strftime('%d')}/{filename}"